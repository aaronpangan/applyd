"""Excel (.xlsx) export builder for jobs.

Each export is a single workbook with two sheets: a ``Summary`` sheet and a
``Jobs`` sheet. Active and archived exports share the ``Jobs`` layout but differ
on the ``Summary`` sheet and accent colour:

* **active**   — headline stat cards (total / bookmarked / interviews / offers),
  a full by-status breakdown, and by-priority + by-work-setup sections.
* **archived** — a trimmed summary with just the total and the generated date,
  plus an extra "Archived Date" column on the ``Jobs`` sheet.
"""

from collections import Counter
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as gl
from openpyxl.worksheet.worksheet import Worksheet

from ..models import Job, JobStatus

_THEMES = {
    "active": {
        "accent": "4F46E5",        # indigo
        "accent_soft": "E0E7FF",
        "band": "F1F3FF",
        "subtitle": "Active Jobs Export",
        "note": "Jobs currently in your active pipeline.",
    },
    "archived": {
        "accent": "B45309",        # burnt amber
        "accent_soft": "FBE8CE",
        "band": "FBF2E4",
        "subtitle": "Archived Jobs Export",
        "note": "Archived jobs — removed from your active list but kept for history.",
    },
}

_STATUS_ORDER = [s.value for s in JobStatus]
_PRIORITY_ORDER = ["high", "medium", "low"]
_WORK_SETUP_ORDER = ["remote", "hybrid", "onsite"]

_THIN = Side(style="thin", color="D4D4D8")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# (header, width, accessor, number_format, alignment)
_BASE_COLUMNS = [
    ("Company", 22, lambda j: j.company, None, "left"),
    ("Position", 26, lambda j: j.position, None, "left"),
    ("Status", 16, lambda j: _pretty(j.status.value), None, "left"),
    ("Priority", 11, lambda j: _pretty(j.priority.value), None, "left"),
    ("Work Setup", 12, lambda j: _pretty(j.work_setup.value), None, "left"),
    ("Location", 18, lambda j: j.location or "", None, "left"),
    ("Industry", 16, lambda j: j.industry or "", None, "left"),
    ("Salary Min", 12, lambda j: j.salary_min, "#,##0", "right"),
    ("Salary Max", 12, lambda j: j.salary_max, "#,##0", "right"),
    ("Currency", 10, lambda j: j.salary_currency, None, "left"),
    ("Recruiter Email", 24, lambda j: j.email_contact or "", None, "left"),
    ("Link", 32, lambda j: j.link, None, "left"),
    ("Job Description", 40, lambda j: j.job_description or "", None, "left"),
    ("Applied Date", 14, lambda j: j.applied_at, "yyyy-mm-dd", "left"),
    ("Created Date", 14, lambda j: _naive(j.created_at), "yyyy-mm-dd", "left"),
]
_ARCHIVED_COLUMN = (
    "Archived Date", 14, lambda j: _naive(j.archived_at), "yyyy-mm-dd", "left",
)


def _pretty(value: str) -> str:
    return value.replace("_", " ").title()


def _naive(value: datetime | None) -> datetime | None:
    """Strip tzinfo — Excel cannot represent timezone-aware datetimes."""
    return value.replace(tzinfo=None) if value is not None else None


def _fill_block(ws: Worksheet, r1, c1, r2, c2, fill=None, border=True) -> None:
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=col)
            if fill:
                cell.fill = fill
            if border:
                cell.border = _BORDER


def _title_block(ws: Worksheet, theme: dict, generated: datetime, span: int) -> None:
    last = gl(span)
    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = "Applyd"
    ws["A1"].font = Font(name="Calibri", bold=True, size=30, color=theme["accent"])
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 42

    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = theme["subtitle"]
    ws["A2"].font = Font(name="Calibri", size=15, bold=True, color="3F3F46")
    ws.row_dimensions[2].height = 22

    ws.merge_cells(f"A3:{last}3")
    ws["A3"] = f"Generated {generated:%B %d, %Y · %H:%M} UTC"
    ws["A3"].font = Font(name="Calibri", size=10, italic=True, color="71717A")

    ws.merge_cells(f"A4:{last}4")
    ws["A4"] = theme["note"]
    ws["A4"].font = Font(name="Calibri", size=10, italic=True, color="52525B")
    ws.row_dimensions[4].height = 18


def _card(ws: Worksheet, col: int, top: int, span: int, number, label, theme) -> None:
    """A stat card: a big number stacked over a label, soft-filled and bordered."""
    c1, c2 = gl(col), gl(col + span - 1)
    ws.merge_cells(f"{c1}{top}:{c2}{top + 2}")
    ws.merge_cells(f"{c1}{top + 3}:{c2}{top + 3}")

    num = ws.cell(row=top, column=col, value=number)
    num.font = Font(name="Calibri", bold=True, size=30, color=theme["accent"])
    num.alignment = Alignment(horizontal="center", vertical="center")

    lbl = ws.cell(row=top + 3, column=col, value=label.upper())
    lbl.font = Font(name="Calibri", bold=True, size=9, color="6B7280")
    lbl.alignment = Alignment(horizontal="center", vertical="center")

    _fill_block(ws, top, col, top + 3, col + span - 1,
                fill=PatternFill("solid", fgColor=theme["accent_soft"]))
    for row in (top, top + 1, top + 2):
        ws.row_dimensions[row].height = 24
    ws.row_dimensions[top + 3].height = 20


def _section(ws: Worksheet, row: int, c1: int, c2: int, label: str, theme) -> None:
    ws.merge_cells(f"{gl(c1)}{row}:{gl(c2)}{row}")
    cell = ws.cell(row=row, column=c1, value=label.upper())
    cell.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    _fill_block(ws, row, c1, row, c2,
                fill=PatternFill("solid", fgColor=theme["accent"]), border=False)
    ws.row_dimensions[row].height = 26


def _kv(ws: Worksheet, row, kc1, kc2, vc1, vc2, key, val, theme) -> None:
    ws.merge_cells(f"{gl(kc1)}{row}:{gl(kc2)}{row}")
    ws.merge_cells(f"{gl(vc1)}{row}:{gl(vc2)}{row}")

    k = ws.cell(row=row, column=kc1, value=key)
    k.font = Font(name="Calibri", size=11, color="3F3F46")
    k.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    v = ws.cell(row=row, column=vc1, value=val)
    v.font = Font(name="Calibri", size=13, bold=True, color=theme["accent"])
    v.alignment = Alignment(horizontal="center", vertical="center")

    _fill_block(ws, row, kc1, row, kc2,
                fill=PatternFill("solid", fgColor=theme["band"]))
    _fill_block(ws, row, vc1, row, vc2)
    ws.row_dimensions[row].height = 24


def _summary_active(ws: Worksheet, jobs: list[Job], theme, generated) -> None:
    for idx in range(1, 9):
        ws.column_dimensions[gl(idx)].width = 13
    _title_block(ws, theme, generated, span=8)

    status = Counter(j.status.value for j in jobs)
    priority = Counter(j.priority.value for j in jobs)
    setup = Counter(j.work_setup.value for j in jobs)
    interviews = (
        status["phone_screen"] + status["tech_interview"] + status["final_interview"]
    )
    offers = status["offer"] + status["accepted"]

    # Headline stat cards.
    _card(ws, 1, 6, 2, len(jobs), "Total Active", theme)
    _card(ws, 3, 6, 2, status["bookmarked"], "Bookmarked", theme)
    _card(ws, 5, 6, 2, interviews, "Interviews", theme)
    _card(ws, 7, 6, 2, offers, "Offers", theme)

    # Full status breakdown.
    _section(ws, 11, 1, 8, "By Status", theme)
    row = 12
    for name in _STATUS_ORDER:
        _kv(ws, row, 1, 5, 6, 8, _pretty(name), status[name], theme)
        row += 1

    # Priority and work setup, side by side.
    head = row + 1
    _section(ws, head, 1, 4, "By Priority", theme)
    _section(ws, head, 5, 8, "By Work Setup", theme)
    for i in range(3):
        r = head + 1 + i
        pr = _PRIORITY_ORDER[i]
        wsx = _WORK_SETUP_ORDER[i]
        _kv(ws, r, 1, 3, 4, 4, _pretty(pr), priority[pr], theme)
        _kv(ws, r, 5, 7, 8, 8, _pretty(wsx), setup[wsx], theme)


def _summary_archived(ws: Worksheet, jobs: list[Job], theme, generated) -> None:
    for idx in range(1, 9):
        ws.column_dimensions[gl(idx)].width = 13
    _title_block(ws, theme, generated, span=8)

    # Archived exports intentionally show only the total and the generated date.
    _card(ws, 1, 6, 4, len(jobs), "Total Archived Jobs", theme)
    _card(ws, 5, 6, 4, f"{generated:%b %d, %Y}", "Generated", theme)


def _jobs_sheet(ws: Worksheet, jobs: list[Job], columns: list, theme) -> None:
    for idx, (_, width, *_rest) in enumerate(columns, start=1):
        ws.column_dimensions[gl(idx)].width = width

    last_col = gl(len(columns))
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"{theme['subtitle']}  —  {len(jobs)} records"
    ws["A1"].font = Font(name="Calibri", bold=True, size=13, color=theme["accent"])
    ws["A1"].fill = PatternFill("solid", fgColor=theme["accent_soft"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    for col_idx, (header, *_rest) in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=theme["accent"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    band = PatternFill("solid", fgColor=theme["band"])
    for r_idx, job in enumerate(jobs, start=3):
        row_fill = band if r_idx % 2 else None
        for c_idx, (header, _w, accessor, number_format, align) in enumerate(
            columns, start=1
        ):
            cell = ws.cell(row=r_idx, column=c_idx, value=accessor(job))
            cell.font = Font(name="Calibri", size=10, color="27272A")
            cell.border = _BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=header == "Job Description",
                horizontal=align,
            )
            if number_format:
                cell.number_format = number_format
            if row_fill:
                cell.fill = row_fill

    ws.auto_filter.ref = f"A2:{last_col}{len(jobs) + 2}"


def build_export_workbook(jobs: list[Job], *, is_archived: bool) -> Workbook:
    """Build the in-memory workbook for an active or archived jobs export."""
    variant = "archived" if is_archived else "active"
    theme = _THEMES[variant]
    generated = datetime.now(timezone.utc)
    columns = _BASE_COLUMNS + ([_ARCHIVED_COLUMN] if is_archived else [])

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.sheet_view.showGridLines = False
    if is_archived:
        _summary_archived(summary, jobs, theme, generated)
    else:
        _summary_active(summary, jobs, theme, generated)

    jobs_sheet = wb.create_sheet("Jobs")
    jobs_sheet.sheet_view.showGridLines = False
    _jobs_sheet(jobs_sheet, jobs, columns, theme)
    return wb
